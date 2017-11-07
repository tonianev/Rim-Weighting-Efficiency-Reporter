"""
	- Script assumes the quarter being generated is within the same year.(Thus Q4 reports must be generated between Dec 1 and Dec 31)
"""
import datetime
import openpyxl
import zipfile
import shutil
import os

#Get current date to be used for comparison with InforSys filename date
cur_date = datetime.datetime.now().strftime ("%Y%m%d")
cur_year = cur_date [2:4]


class WeightingEfficiency:

	def __init__(self):
		self.selected_q = []
		self.quarter_name = ''
		self.target_dir = ''

	def user_input(self):

		first_q  = ['01','02','03']
		second_q = ['04','05','06']
		third_q  = ['07','08','09']
		fourth_q = ['10','11','12']

		usr_q = int(input("Which quarter are you looking for? "))
		
		if usr_q == 1:
			self.selected_q = first_q
			self.quarter_name = 'Q1' 
		elif usr_q == 2:
			self.selected_q = second_q
			self.quarter_name = 'Q2' 
		elif usr_q == 3:
			self.selected_q = third_q
			self.quarter_name = 'Q3' 
		elif usr_q == 4:
			self.selected_q = fourth_q
			self.quarter_name = 'Q4' 
		else:
			print("Please enter '1' for Q1, '2' for Q2, '3' for Q3 or '4' for Q4")

	def make_target_dir(self):
		# Check if target folder exists and if not create one.

		self.target_dir = "xxx.xxx.xx.xx/Reports_Dir/Weight_Report_" + cur_date[0:4] + "_" + self.quarter_name

		if not os.path.exists(self.target_dir):
			os.mkdir(self.target_dir)
		else:
			print("Looks like you might already have this report.")

	def transfer_files(self):
		"""Take target files from InfoSys and copy them on local folder"""

		host_drive = "xxx.xxx.xx.xx"

		for filename in os.listdir(host_drive):
		    if filename.startswith("XXXX_XXXX" + cur_year + self.selected_q[0] + '01') or filename.startswith("XXXX_XXXX" + cur_year + self.selected_q[0] + '01') and filename.endswith(".zip"): 
		    	shutil.copy(os.path.join(host_drive, filename), self.target_dir)

		    elif filename.startswith("XXXX_XXXX" + cur_year + self.selected_q[1] + '01') or filename.startswith("XXXX_XXXX" + cur_year + self.selected_q[1] + '01') and filename.endswith(".zip"): 
		    	shutil.copy(os.path.join(host_drive, filename), self.target_dir)

		    elif filename.startswith("XXXX_XXXX" + cur_year + self.selected_q[2] + '01') or filename.startswith("XXXX_XXXX" + cur_year + self.selected_q[2] + '01') and filename.endswith(".zip"): 
		    	shutil.copy(os.path.join(host_drive, filename), self.target_dir)
	    
	def sort_files(self):
		"""Sort copied PT files into TV and Radio Folders"""
		try:
			for file in os.listdir(self.target_dir):
				if file.startswith("XXXX_XXXX") and file.endswith(".zip"):
					zf = zipfile.ZipFile(os.path.join(self.target_dir, file), 'r')
					for filename in zf.namelist():
						if filename.endswith(".DEM"):
							zf.extract(filename, path = os.path.join(self.target_dir,'TV_Data'))

				elif file.startswith("XXXX_XXXX") and file.endswith(".zip"):
					zf = zipfile.ZipFile(os.path.join(self.target_dir, file), 'r')
					for filename in zf.namelist():
						if filename.endswith(".DEM"):
							zf.extract(filename, path = os.path.join(self.target_dir,'Radio_Data'))
		finally:
			zf.close()
			for file in os.listdir(self.target_dir):
				if file.endswith(".zip"):
					os.remove(os.path.join(self.target_dir, file))

	def dem_parse(self, dem_file):

		""" - Opens .DEM file and targets "."
			- List is parsed to only contain home number and memeber number
		"""
		with open(dem_file) as f:
			read_data = []
			file_name = f.name[-12:]
			for line in f:
				read_data.append(line)
		
		home_no = [i[0:7] for i in read_data]
		mem_no = [i[7:9] for i in read_data]
		weight_no = [i[9:17] for i in read_data]
		panel_char = [i[17:18] for i in read_data]
		demo_33 = [i[32:33] for i in read_data]
		demo_34 = [i[33:34] for i in read_data]
		demo_68 = [i[67:68] for i in read_data]
		demo_90 = [i[89:90] for i in read_data]
		demo_99 = [i[98:99] for i in read_data]
		demo_226 = [i[225:226] for i in read_data]

		return file_name, home_no, mem_no, weight_no, panel_char, demo_33, demo_34, demo_68, demo_90, demo_99, demo_226

	def write_excel(self):

		tv_dems = self.target_dir + "/TV_Data/"
		radio_dems = self.target_dir + "/Radio_Data/"

		wb = openpyxl.Workbook()

		for file in os.listdir(tv_dems):
			file_name, home_no, mem_no, weight_no, panel_char, demo_33, demo_34, demo_68, demo_90, demo_99, demo_226 = self.dem_parse(os.path.join(tv_dems,file))
			ws1 = wb.create_sheet('TV_{}'.format(file_name), 0)
			ws1.append(["home", "mem", "weight", "panel", "33", "34", "68", "99", "226"])
			for row in zip(home_no, mem_no, weight_no, panel_char, demo_33, demo_34, demo_68, demo_99, demo_226):
				ws1.append(row) 

		for file in os.listdir(radio_dems):
			file_name, home_no, mem_no, weight_no, panel_char, demo_33, demo_34, demo_68, demo_90, demo_99, demo_226 = self.dem_parse(os.path.join(radio_dems,file))
			ws2 = wb.create_sheet('Radio_{}'.format(file_name), 3)
			ws2.append(["home", "mem", "weight", "panel", "33", "34", "90", "99", "226"])
			for row in zip(home_no, mem_no, weight_no, panel_char, demo_33, demo_34, demo_90, demo_99, demo_226):
				ws2.append(row) 

		os.chdir(self.target_dir)
		wb.save('weighting_efficiency_{0}_{1}.xlsx'.format(cur_date[0:4], self.quarter_name))


if __name__ == '__main__':

	p = WeightingEfficiency()

	try:

		p.user_input()
		p.make_target_dir()
		p.transfer_files()
		p.sort_files()
		p.write_excel()
	
	finally:
		print("All done.")
		input("Press enter to exit...")
